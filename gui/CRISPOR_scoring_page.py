from pathlib import Path
import os

from PySide6.QtWidgets import (
    QWidget,
    QMessageBox,
    QVBoxLayout,
    QLineEdit,
    QLabel,
    QGroupBox,
    QHBoxLayout,
    QPushButton,
    QFileDialog,
    QTableWidget,
    QTableWidgetItem,
    QAbstractItemView,
    QProgressBar
)

from PySide6.QtCore import(
    Qt,
    Signal,
    QThread
)
from PySide6.QtGui import (
    QColor,
    QPixmap,
)
import openpyxl
from openpyxl.drawing.image import Image as XLImage


from core.crispor_analyser import AnalyzerMachine
from core.config_manager import load_configs

COL_ID = 0
COL_SEQ = 1
COL_PAM = 2
COL_DOENCH = 3
COL_OUR_TOTAL_SCORE = 4
COL_RESTR = 5
COL_PIC = 6

COLUMN_HEADERS = ['id', 'Sequence', 'PAM', 'Doench','OTS', 'Restrictase','Picture']

COLOR_LOW_DOENCH = QColor(243, 139, 168).darker(160)   #or bad
COLOR_MID_DOENCH = QColor(249, 226, 175).darker(160)   
COLOR_HIGH_DOENCH = QColor(148, 226, 161).darker(160)
COLOR_HIGHIEST_DOENCH = QColor(166, 227, 161).darker(140)



class CrisporAnalyzerPage(QWidget):
    configs_changed = Signal(dict)

    def __init__(self, parent = None):
        super().__init__(parent)
        self.results: list[GRNARecord] = []
        self._machinery: AnalysisWorker | None = None
        self._thread: QThread | None = None
        self._build_gui()

    def _build_gui(self):
        self.root_layout = QVBoxLayout()
        self.crispor_settings = QGroupBox('Настройки анаиза выдачи CRISPOR')
        self.crispor_settings_layout = QVBoxLayout(self.crispor_settings) 
        self.row1 = QHBoxLayout(self.crispor_settings)
        self.file_path_browser_editline = QLineEdit()
        self.file_path_browser_editline.setPlaceholderText('Путь к файлу...')

        self.file_path_browser_button = QPushButton()
        self.file_path_browser_button.setText('Найти файл')
        self.file_path_browser_button.clicked.connect(self._browse_file)
        
        self.start_analysis_button = QPushButton()
        self.start_analysis_button.setText('▶  Анализ')
        self.start_analysis_button.clicked.connect(self.start_analysis)

        self.row1.addWidget(QLabel('Файл:'))
        self.row1.addWidget(self.file_path_browser_editline)
        self.row1.addWidget(self.file_path_browser_button)
        self.row1.addWidget(self.start_analysis_button)

        self.row2 = QHBoxLayout(self.crispor_settings)
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)   
        self.progress_bar.setVisible(False)
        self.status_label = QLabel('')
        self.status_label.setObjectName('status_label')
        self.status_label.setVisible(False)
        self.text_progress = QLabel('Прогресс:')
        self.text_progress.setVisible(False)
        self.row2.addStretch()
        self.row2.addWidget(self.text_progress)
        self.row2.addWidget(self.progress_bar, 3)
        self.row2.addWidget(self.status_label, 2)
        self.row2.addStretch()

        self.crispor_settings_layout.addLayout(self.row1)
        self.crispor_settings_layout.addLayout(self.row2)

        self.crispor_table = QTableWidget(0,len(COLUMN_HEADERS))
        self.crispor_table.setHorizontalHeaderLabels(COLUMN_HEADERS)
        self.crispor_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.crispor_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.crispor_table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        self.crispor_table.setSortingEnabled(True)
        self.crispor_table.resizeColumnsToContents()



        self.control_panel = QGroupBox()
        self.control__panel_row = QHBoxLayout(self.control_panel)
        self.delete_chosen_elements_button = QPushButton()
        self.delete_chosen_elements_button.setText('Удалить выбранное')
        self.delete_chosen_elements_button.setProperty('class','delete')
        self.delete_chosen_elements_button.clicked.connect(self._delete_selected_rows)
        self.export_to_excel_button = QPushButton()
        self.export_to_excel_button.setText('Экспорт в эксель')
        self.export_to_excel_button.clicked.connect(self._export_excel)
        self.delete_all_bad_button = QPushButton()
        self.delete_all_bad_button.setText('Удалить все плохие')
        self.delete_all_bad_button.setProperty('class','delete')
        self.delete_all_bad_button.clicked.connect(self._delete_all_bad_rows)
        self.control__panel_row.addWidget(self.delete_chosen_elements_button)
        self.control__panel_row.addWidget(self.delete_all_bad_button)
        self.control__panel_row.addWidget(self.export_to_excel_button)
        self.control__panel_row.addStretch()

        self.root_layout.addWidget(self.crispor_settings)
        self.root_layout.addWidget(self.crispor_table)
        self.root_layout.addWidget(self.control_panel)

        self.setLayout(self.root_layout)

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, 'Выбери файл CRISPOR', '',
            'Excel файлы (*.xlsx *.xls);;Все файлы (*)'
        )
        if path:
            self.file_path_browser_editline.setText(path)
        print(path)
    
    def _delete_selected_rows(self):
        selected_rows = sorted(
            set(item.row() for item in self.crispor_table.selectedItems()),
            reverse=True,
        )
        for row in selected_rows:
            self.crispor_table.removeRow(row)
        print('deleting...')

    def _delete_all_bad_rows(self):
        rows_to_delete = []
        for row in range(self.crispor_table.rowCount()):
            doench_item = self.crispor_table.item(row, COL_DOENCH)
            pam_item = self.crispor_table.item(row, COL_PAM)
            sequence_item = self.crispor_table.item(row, COL_SEQ)

            is_bad = False
            if doench_item and doench_item.background().color() == COLOR_LOW_DOENCH:
                is_bad = True
            if pam_item and pam_item.background().color() == COLOR_LOW_DOENCH:
                is_bad = True
            if sequence_item and sequence_item.background().color() == COLOR_LOW_DOENCH:
                is_bad = True

            if is_bad:
                rows_to_delete.append(row)

        # Удаляем снизу вверх — иначе индексы съедут
        for row in sorted(rows_to_delete, reverse=True):
            self.crispor_table.removeRow(row)
        

    def start_analysis(self):
        self.path_crispor = self.file_path_browser_editline.text()
        if not self.path_crispor:
            QMessageBox.warning(self,'Ошибка', 'Выберите сначала файл')
            return
        if not Path(self.path_crispor).exists():
            QMessageBox.warning(self, 'Ошибка', 'Выбранный файл не существует')
            return

        if load_configs('preset_gRNA_tails', load_configs('main_configs','selected_tail')) != None:
            self.tail = load_configs('preset_gRNA_tails', load_configs('main_configs','selected_tail'))
        elif load_configs('custom_gRNA_tails', load_configs('main_configs','selected_tail')) != None:
            self.tail = load_configs('custom_gRNA_tails', load_configs('main_configs','selected_tail'))

        self.file_name = os.path.basename(self.path_crispor) #преобразование ссылки в название файла
        self.file_name = os.path.splitext(self.file_name)[0]
        self.output_dir = Path(__file__).parent.parent / f'output/{self.file_name}'


        self.crispor_table.setRowCount(0)
        self.results.clear()
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)
        self.text_progress.setVisible(True)
        self.status_label.setVisible(True)
        self.delete_chosen_elements_button.setEnabled(False)
        self.export_to_excel_button.setEnabled(False)


        if self._thread is not None and self._thread.isRunning():
            if self._machinery is not None:
                self._machinery.cancel()
            self._thread.quit()
            self._thread.wait(30)
        if self._thread is not None:
            self._thread.deleteLater()
            self._thread = None
        if self._machinery is not None:
            self._machinery.deleteLater()
            self._machinery = None

        self._thread = QThread()
        self._machinery = AnalyzerMachine(
            crispor_path= self.path_crispor,
            grna_tail= self.tail,
            output_dir= self.output_dir,
            configs=None,
            doench_border=50,
        )
        self._machinery.moveToThread(self._thread)
        self._thread.started.connect(self._machinery.try_run)
        self._machinery.progress.connect(self.progress_bar.setValue)
        self._machinery.status_message.connect(self.status_label.setText)
        self._machinery.result_ready.connect(self._analysis_done)
        self._thread.start()

        
        print('analysis...')
    
    def _analysis_done(self,results):
        self.results = results
        self._populate_tabel(results)
        self.progress_bar.setVisible(False)
        self.status_label.setVisible(False)
        self.text_progress.setVisible(False)
        self.start_analysis_button.setEnabled(True)
        self.delete_chosen_elements_button.setEnabled(True)
        self.export_to_excel_button.setEnabled(True)


    def _populate_tabel(self,analysis_result):
        self.crispor_table.setSortingEnabled(False)
        self.crispor_table.setRowCount(len(analysis_result))

        for row_index, grna_data in enumerate(analysis_result):
            index_item = QTableWidgetItem()
            index_item.setData(Qt.EditRole, row_index + 1)

            doench_item = QTableWidgetItem()
            doench_item.setData(Qt.EditRole, grna_data.doench_16)

            our_total_score_item = QTableWidgetItem()
            our_total_score_item.setData(Qt.EditRole,grna_data.total_score)

            rna_picture = QTableWidgetItem()
            rna_picture.setTextAlignment(Qt.AlignCenter)

            rna_picture_path =  grna_data.png_2d_path

            rna_pixmap = QPixmap(str(rna_picture_path))
            scaled_rna = rna_pixmap.scaled(
                256,256,
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            rna_picture.setData(Qt.DecorationRole, scaled_rna)


            items = [
                index_item,
                QTableWidgetItem(grna_data.sequence),
                QTableWidgetItem(grna_data.PAM),
                doench_item,
                our_total_score_item,
                QTableWidgetItem(', '.join(grna_data.restriction_sites) if grna_data.restriction_sites else ''),
                rna_picture
            ]

            for col, item in enumerate(items):
                item.setTextAlignment(Qt.AlignCenter)
                if col == COL_SEQ:
                    item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                if col == COL_RESTR:
                    item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                if col == COL_PAM:
                    item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                if col == COL_PIC:
                    item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                self.crispor_table.setItem(row_index, col, item)

                        # Подсветка строк
            if int(grna_data.doench_16) < 50:
                self.crispor_table.item(row_index, COL_DOENCH).setBackground(COLOR_LOW_DOENCH)
            elif int(grna_data.doench_16) < 60:
                self.crispor_table.item(row_index, COL_DOENCH).setBackground(COLOR_MID_DOENCH)
            elif int(grna_data.doench_16) < 70:
                self.crispor_table.item(row_index, COL_DOENCH).setBackground(COLOR_HIGH_DOENCH)
            elif int(grna_data.doench_16) < 80:
                self.crispor_table.item(row_index, COL_DOENCH).setBackground(COLOR_HIGHIEST_DOENCH)

            if grna_data.PAM == 'TGG':
                self.crispor_table.item(row_index, COL_PAM).setBackground(COLOR_LOW_DOENCH)

            if grna_data.oligoT == True:
                self.crispor_table.item(row_index, COL_SEQ).setBackground(COLOR_LOW_DOENCH)
            if grna_data.oligoG == True:
                self.crispor_table.item(row_index, COL_SEQ).setBackground(COLOR_LOW_DOENCH)
            # if grna_data.oligoC == True:
            #     self.crispor_table.item(row_index, COL_SEQ).setBackground(COLOR_LOW_DOENCH)
            if grna_data.mfe_bad == True:
                self.crispor_table.item(row_index, COL_SEQ).setBackground(COLOR_LOW_DOENCH)
            if grna_data.eight_links == True:
                self.crispor_table.item(row_index, COL_SEQ).setBackground(COLOR_LOW_DOENCH)
            if grna_data.twelve_links == True:
                self.crispor_table.item(row_index, COL_SEQ).setBackground(COLOR_LOW_DOENCH)
            if grna_data.GCC_not_at_16to20 == True:
                self.crispor_table.item(row_index, COL_SEQ).setBackground(COLOR_LOW_DOENCH)            

            # Сохраняем индекс записи в первой ячейке (для доступа к GRNARecord)
            self.crispor_table.item(row_index, COL_ID).setData(Qt.UserRole, row_index)
            self.crispor_table.setColumnWidth(COL_PIC, 256+8)
            self.crispor_table.setColumnWidth(COL_RESTR, 256+8)
            self.crispor_table.setRowHeight(row_index, 256 + 8)

        self.crispor_table.setSortingEnabled(True)
        self.crispor_table.sortItems(COL_DOENCH, Qt.DescendingOrder)
        self.crispor_table.resizeColumnsToContents()
        self.crispor_table.setColumnWidth(COL_RESTR, 256+8)
        self.crispor_table.setRowHeight(row_index, 256 + 8)


    def _export_excel(self):
        """Экспортирует оставшиеся в таблице гайды в xlsx с PNG структур."""
        if self.crispor_table.rowCount() == 0:
            QMessageBox.information(self, 'Экспорт', 'Таблица пуста.')
            return

        save_path, _ = QFileDialog.getSaveFileName(
            self, 'Сохранить Excel', '', 'Excel файлы (*.xlsx)'
        )
        if not save_path:
            return
        if not save_path.endswith('.xlsx'):
            save_path += '.xlsx'

        try:
            self._write_excel(save_path)
            QMessageBox.information(
                self, 'Экспорт завершён',
                f'Файл сохранён:\n{save_path}'
            )
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка экспорта', str(e))


    def _write_excel(self, save_path: str):
        """Записывает xlsx с данными и PNG вторичных структур."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'gRNA analysis'

        # Заголовки
        headers = COLUMN_HEADERS
        ws.append(headers)

        # Ширина колонок
        col_widths = [5, 25, 12, 8, 14, 6, 12, 40, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Данные
        for row_idx in range(self.crispor_table.rowCount()):
            num_item = self.crispor_table.item(row_idx, COL_ID)
            rec_idx = num_item.data(Qt.UserRole) if num_item else None

            row_data = []
            for col in range(len(COLUMN_HEADERS)):
                item = self.crispor_table.item(row_idx, col)
                row_data.append(item.text() if item else '')
            row_data.append('')  # placeholder для структуры

            xl_row = row_idx + 2  # +2 из-за заголовка
            ws.append(row_data)
            ws.row_dimensions[xl_row].height = 80

            # Вставляем PNG структуры
            if rec_idx is not None and rec_idx < len(self.results):
                rec = self.results[rec_idx]
                if rec.png_2d_path and os.path.exists(rec.png_2d_path):
                    try:
                        img = XLImage(rec.png_2d_path)
                        img.width = 100
                        img.height = 100
                        struct_col = openpyxl.utils.get_column_letter(len(COLUMN_HEADERS) + 1)
                        ws.add_image(img, f'{struct_col}{xl_row}')
                    except Exception:
                        pass

        wb.save(save_path)

