import os

from PySide6.QtWidgets import (
    QWidget,
    QMessageBox,
    QVBoxLayout,
    QGroupBox,
    QHBoxLayout,
    QPushButton,
    QFileDialog,
    QTableWidget,
    QTableWidgetItem,
    QAbstractItemView,
)

from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QColor, QPixmap

import openpyxl
from openpyxl.drawing.image import Image as XLImage


COL_ID = 0
COL_SEQ = 1
COL_PAM = 2
COL_DOENCH = 3
COL_OUR_TOTAL_SCORE = 4
COL_RESTR = 5
COL_PIC = 6
COL_DISC = 7

COLUMN_HEADERS = ['id', 'Sequence', 'PAM', 'Doench', 'OTS', 'Restrictase, 100/250/500 bp', 'Picture','Disadvantages']

COLOR_LOW_DOENCH = QColor(243, 139, 168).darker(160)
COLOR_MID_DOENCH = QColor(249, 226, 175).darker(160)
COLOR_HIGH_DOENCH = QColor(148, 226, 161).darker(160)
COLOR_HIGHIEST_DOENCH = QColor(166, 227, 161).darker(140)


class GrnaTableWidget(QWidget):

    guide_selected = Signal(int)      
    guides_deleted = Signal(list)    
    pair_mode_toggled = Signal(bool) 
    pair_filter_reset = Signal()     

    def __init__(self, parent=None):
        super().__init__(parent)
        self.results: list = []
        self._build_gui()

    def _build_gui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)

        self.table = QTableWidget(0, len(COLUMN_HEADERS))
        self.table.setHorizontalHeaderLabels(COLUMN_HEADERS)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSortingEnabled(True)
        self.table.resizeColumnsToContents()

        self.control_panel = QGroupBox()
        control_layout = QHBoxLayout(self.control_panel)

        self.delete_chosen_button = QPushButton('Удалить выбранное')
        self.delete_chosen_button.setProperty('class', 'delete')
        self.delete_chosen_button.clicked.connect(self._delete_selected_rows)

        self.delete_all_bad_button = QPushButton('Удалить все плохие')
        self.delete_all_bad_button.setProperty('class', 'delete')
        self.delete_all_bad_button.clicked.connect(self._delete_all_bad_rows)

        self.export_to_excel_button = QPushButton('Экспорт в эксель')
        self.export_to_excel_button.clicked.connect(self._export_excel)

        self.pair_mode_button = QPushButton('Подбор пары')
        self.pair_mode_button.setCheckable(True)
        self.pair_mode_button.toggled.connect(lambda checked: self.pair_mode_toggled.emit(checked))
        self.pair_mode_button.setProperty('class','pairs')

        self.pair_reset_button = QPushButton('Сбросить фильтр')
        self.pair_reset_button.setVisible(False)
        self.pair_reset_button.clicked.connect(self.pair_filter_reset)
        self.pair_reset_button.setProperty('class','pairs')

        control_layout.addWidget(self.delete_chosen_button)
        control_layout.addWidget(self.delete_all_bad_button)
        control_layout.addWidget(self.export_to_excel_button)
        control_layout.addWidget(self.pair_mode_button)
        control_layout.addWidget(self.pair_reset_button)
        control_layout.addStretch()

        root.addWidget(self.table)
        root.addWidget(self.control_panel)

        self.table.itemSelectionChanged.connect(self._on_selection_changed)


    def populate(self, results: list):
        self.results = results
        self._populate_table(results)

    def clear(self):
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        self.results.clear()
        self.table.setSortingEnabled(True)

    def set_controls_enabled(self, enabled: bool):
        self.delete_chosen_button.setEnabled(enabled)
        self.delete_all_bad_button.setEnabled(enabled)
        self.export_to_excel_button.setEnabled(enabled)
        self.pair_mode_button.setEnabled(enabled)

    def _on_selection_changed(self):
        rows = set(item.row() for item in self.table.selectedItems())
        if not rows:
            return
        row = min(rows)
        item = self.table.item(row, COL_ID)
        if item is not None:
            idx = item.data(Qt.UserRole)
            if idx is not None:
                self.guide_selected.emit(idx)

    def select_guide_by_index(self, idx: int):
        for row in range(self.table.rowCount()):
            item = self.table.item(row, COL_ID)
            if item and item.data(Qt.UserRole) == idx:
                self.table.selectRow(row)
                self.table.scrollToItem(self.table.item(row, 0))
                return

    def _populate_table(self, analysis_result):
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(analysis_result))

        for row_index, grna_data in enumerate(analysis_result):
            index_item = QTableWidgetItem()
            index_item.setData(Qt.EditRole, grna_data.indx)

            doench_item = QTableWidgetItem()
            doench_item.setData(Qt.EditRole, grna_data.doench_16)

            our_total_score_item = QTableWidgetItem()
            our_total_score_item.setData(Qt.EditRole, grna_data.total_score)

            rna_picture = QTableWidgetItem()
            rna_picture.setTextAlignment(Qt.AlignCenter)

            rna_picture_path = grna_data.png_2d_path
            rna_pixmap = QPixmap(str(rna_picture_path))
            scaled_rna = rna_pixmap.scaled(
                256, 256,
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation,
            )
            rna_picture.setData(Qt.DecorationRole, scaled_rna)

            disadvantages_list = ''
            if int(grna_data.doench_16) < 50:
                disadvantages_list += 'low Doench 16\n'
            if grna_data.PAM == 'TGG':
                disadvantages_list += 'TGG in PAM\n'
            if grna_data.oligoT:
                disadvantages_list += 'oligoT\n'
            if grna_data.mfe_bad:
                disadvantages_list += 'low MFE\n'
            if grna_data.eight_links:
                disadvantages_list += 'eight links in a row\n'
            if grna_data.twelve_links:
                disadvantages_list += 'twelve links in total\n'
            if grna_data.GCC_not_at_16to20:
                disadvantages_list += 'GCC at 16-20pos\n'
            if not grna_data.dot_bracked_structure.endswith('((((....))))(((((((...)))))))...'):
                disadvantages_list += 'wrong harpins\n'
                
            disadvantages_item = QTableWidgetItem()
            disadvantages_item.setData(Qt.EditRole, disadvantages_list)

            items = [
                index_item,
                QTableWidgetItem(grna_data.sequence),
                QTableWidgetItem(grna_data.PAM),
                doench_item,
                our_total_score_item,
                QTableWidgetItem(', '.join(grna_data.restriction_sites) if grna_data.restriction_sites else ''),
                rna_picture,
                disadvantages_item
            ]

            for col, item in enumerate(items):
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_index, col, item)

            self._highlight_row(row_index, grna_data)

            self.table.item(row_index, COL_ID).setData(Qt.UserRole, grna_data.indx)
            self.table.setColumnWidth(COL_PIC, 256 + 8)
            self.table.setColumnWidth(COL_RESTR, 256 + 8)
            self.table.setRowHeight(row_index, 256 + 8)

        self.table.setSortingEnabled(True)
        self.table.sortItems(COL_DOENCH, Qt.DescendingOrder)
        self.table.resizeColumnsToContents()
        self.table.setColumnWidth(COL_RESTR, 256 + 8)
        if self.table.rowCount() > 0:
            self.table.setRowHeight(self.table.rowCount() - 1, 256 + 8)

    def _highlight_row(self, row_index, grna_data):
        if int(grna_data.doench_16) < 50:
            self.table.item(row_index, COL_DOENCH).setBackground(COLOR_LOW_DOENCH)
        elif int(grna_data.doench_16) < 60:
            self.table.item(row_index, COL_DOENCH).setBackground(COLOR_MID_DOENCH)
        elif int(grna_data.doench_16) < 70:
            self.table.item(row_index, COL_DOENCH).setBackground(COLOR_HIGH_DOENCH)
        elif int(grna_data.doench_16) < 80:
            self.table.item(row_index, COL_DOENCH).setBackground(COLOR_HIGHIEST_DOENCH)

        if grna_data.PAM == 'TGG':
            self.table.item(row_index, COL_PAM).setBackground(COLOR_LOW_DOENCH)

        if grna_data.oligoT:
            self.table.item(row_index, COL_SEQ).setBackground(COLOR_LOW_DOENCH)
        if grna_data.mfe_bad:
            self.table.item(row_index, COL_SEQ).setBackground(COLOR_LOW_DOENCH)
        if grna_data.eight_links:
            self.table.item(row_index, COL_SEQ).setBackground(COLOR_LOW_DOENCH)
        if grna_data.twelve_links:
            self.table.item(row_index, COL_SEQ).setBackground(COLOR_LOW_DOENCH)
        if grna_data.GCC_not_at_16to20:
            self.table.item(row_index, COL_SEQ).setBackground(COLOR_LOW_DOENCH)
        if not grna_data.dot_bracked_structure.endswith('((((....))))(((((((...)))))))...'):
            self.table.item(row_index, COL_SEQ).setBackground(COLOR_LOW_DOENCH)
            
    def _delete_selected_rows(self):
        selected_rows = sorted(
            set(item.row() for item in self.table.selectedItems()),
            reverse=True,
        )
        if not selected_rows:
            return

        remainings = []
        for row in range(self.table.rowCount()):
            if row not in selected_rows:
                item = self.table.item(row, COL_ID)
                if item:
                    remainings.append(item.data(Qt.UserRole))

        for row in selected_rows:
            self.table.removeRow(row)

        self.guides_deleted.emit(remainings)

    def _delete_all_bad_rows(self):
        rows_to_delete = []
        for row in range(self.table.rowCount()):
            is_bad = False
            for col in (COL_DOENCH, COL_PAM, COL_SEQ):
                item = self.table.item(row, col)
                if item and item.background().color() == COLOR_LOW_DOENCH:
                    is_bad = True
                    break
            if is_bad:
                rows_to_delete.append(row)

        if not rows_to_delete:
            return

        remainings = []
        for row in range(self.table.rowCount()):
            if row not in rows_to_delete:
                item = self.table.item(row, COL_ID)
                if item:
                    remainings.append(item.data(Qt.UserRole))

        for row in sorted(rows_to_delete, reverse=True):
            self.table.removeRow(row)

        self.guides_deleted.emit(remainings)

    def _export_excel(self):
        if self.table.rowCount() == 0:
            QMessageBox.information(self, 'Экспорт', 'Таблица пуста.')
            return

        save_path, _ = QFileDialog.getSaveFileName(
            self, 'Сохранить Excel', '', 'Excel файлы (*.xlsx)',
        )
        if not save_path:
            return
        if not save_path.endswith('.xlsx'):
            save_path += '.xlsx'

        try:
            self._write_excel(save_path)
            QMessageBox.information(self, 'Экспорт завершён', f'Файл сохранён:\n{save_path}')
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка экспорта', str(e))

    def _write_excel(self, save_path: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'gRNA analysis'

        ws.append(COLUMN_HEADERS)

        col_widths = [5, 25, 12, 8, 14, 6, 12, 40, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        for row_idx in range(self.table.rowCount()):
            num_item = self.table.item(row_idx, COL_ID)
            rec_idx = num_item.data(Qt.UserRole) if num_item else None

            row_data = []
            for col in range(len(COLUMN_HEADERS)):
                item = self.table.item(row_idx, col)
                row_data.append(item.text() if item else '')
            row_data.append('')

            xl_row = row_idx + 2
            ws.append(row_data)
            ws.row_dimensions[xl_row].height = 80

            rec = next((r for r in self.results if r.indx == rec_idx), None)
            if rec is not None:
                if rec.png_2d_path and os.path.exists(rec.png_2d_path):
                    try:
                        img = XLImage(rec.png_2d_path)
                        img.width = 100
                        img.height = 100
                        struct_col = openpyxl.utils.get_column_letter(len(COLUMN_HEADERS) + 1)
                        ws.add_image(img, f'{struct_col}{xl_row}')
                    except Exception:
                        pass

        wb.save(save_path)
