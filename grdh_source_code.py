import pandas as pd
from dataclasses import dataclass, field
import customtkinter as CTk
from tkinter import filedialog
from CTkTable import *
import xlrd #необходима для чтения .xls файлов. В норме pandas умеет читать только .xlsx
from ViennaRNA import RNA
import os
import time
import cairosvg
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import csv
import openpyxl

'''
выгрузка данных RNAfold библиотекой ViennaRNA https://viennarna.readthedocs.io/en/latest/getting_started.html

сборка в exe с помощью python -m auto_py_to_exe
'''


@dataclass
class gRNA_sequence():
    sequence : str 
    PAM : str
    oligoT : int = 0
    strain : int = 0
    GC_freq : int = 0 
    GC_count_10to20 : int = 0 
    last_four_pur : int = 0 
    access_18_to_20 : int = 0
    seven_links : int = 0
    twelve_links : int = 0
    C_not_at_3 : int = 0 
    G_not_at_16 : int = 0 
    C_at_16 : int = 0 
    G_or_A_at_20 : int = 0 
    C_at_18 : int = 0 
    G_not_at_14 : int = 0
    GCC_not_at_16to20 : int = 0
    PAMs_N : int = 0
    restriction_sites : list = ()
    total_score : int = 0

class App(CTk.CTk):
    def __init__(self):
        super().__init__()

        self.title('gRNA helper')
        self.geometry('800x400')
        self.resizable(width= False, height= False)

        self.button_to_start = CTk.CTkButton(self, text ='начало анализа',fg_color = 'orange', text_color = 'dark blue', width = 50, height = 20, command = self.start_code_by_button)
        self.button_to_start.grid(row = 0, column = 7,padx=(10,20),pady =(10,20), sticky = "ew" )


        self.input_text_frame = CTk.CTkLabel(master = self,text = 'Путь к файлу:', width= 100, height= 20)
        self.input_text_frame.grid(row = 0, column = 0, padx=(10,20), pady =(10,20), sticky = "ew" )
        self.to_input_file_frame = CTk.CTkEntry(master = self, width= 250, height= 20, placeholder_text = '.xls файл полученный от CRISPOR')
        self.to_input_file_frame.grid(row = 0, column = 1, padx=(10,20),pady =(10,20), sticky = "ew",columnspan = 2 )
        self.browse_file = CTk.CTkButton(self, text ='найти файл',fg_color = 'forest green', text_color = 'black', width = 50, height = 20, command=lambda:self.choose_file(self.to_input_file_frame))
        self.browse_file.grid(row = 0, column = 4,padx=(10,20),pady =(10,20), sticky = "ew" )

        self.configs_text = CTk.CTkLabel(master = self,text = 'Настройки:', width= 100, height= 20)
        self.configs_text.grid(row = 1, column = 0, padx=(10,20), pady =(10,20), sticky = "ew" )
        self.configs_path = CTk.CTkEntry(master = self, width= 250, height= 20, placeholder_text = 'выбраны настройки по умолчанию')
        self.configs_path.grid(row = 1, column = 1, padx=(10,20),pady =(10,20), sticky = "ew",columnspan = 2 )
        self.browse_configs = CTk.CTkButton(self, text ='загрузить настройки',fg_color = 'forest green', text_color = 'black', width = 50, height = 20, command=lambda:self.choose_file(self.configs_path))
        self.browse_configs.grid(row = 1, column = 4,padx=(10,20),pady =(10,20), sticky = "ew" )

        self.input_RNA_tail_frame = CTk.CTkLabel(master = self,text = 'Хвост gRNA:', width= 100, height= 20)
        self.input_RNA_tail_frame.grid(row = 2, column = 0, padx=(10,20), pady =(10,20), sticky = "ew" )
        self.input_RNA_tail_seq_frame = CTk.CTkEntry(master = self, width= 250, height= 20, placeholder_text = 'GTTTCAGAGCTATGCTGGAAACAGCATAGCAAGTTGAAATAAGGCTAGTCCGTTATCAACTTGAAAAAGTGGCACCGAGTCGGTGCTTTT')
        self.input_RNA_tail_seq_frame.grid(row = 2, column = 1, padx=(10,20),pady =(10,20), sticky = "ew",columnspan = 2 )

    def programm_is_done(self,path_to_result):
        notion = CTk.CTkToplevel()
        notion.title("Сообщение")
        notion.geometry("350x150")
        notion.transient(app)
        notion.grab_set()
        label = CTk.CTkLabel(notion, text=f"Программа завершила работу\nРезультаты сохранены в:\n{path_to_result}")
        label.pack(pady=20)
        button = CTk.CTkButton(notion, text="OK", command=notion.destroy)
        button.pack(pady=10)

    def choose_file(self,grid_to_place):
        choosen_file = filedialog.askopenfilename(title="Выберите файл", filetypes=[("Все файлы", "*.*")])
        grid_to_place.delete(0, "end")
        grid_to_place.insert(0, choosen_file)

    def GC_structure(self, sequence):
        gc_c = sequence.count('G') + sequence.count('C')
        gc_f = int(round(gc_c / len(sequence)*100,0))
        gc_c_10_20 = sequence[9:20].count('G') + sequence[9:20].count('C')
        last_four_nuc = sequence[16:20].count('A') + sequence[16:20].count('G')
        return(gc_c_10_20, gc_f,last_four_nuc)

    def RNA_2D_structure(self, sequence,PAM_sequence, gRNA_tail_sequence): #создание svg картинки вторичной структуры gRNA
        sequence = sequence + PAM_sequence + gRNA_tail_sequence
        fc = RNA.fold_compound(sequence)
        structure, energy = fc.mfe()
        
        output_file = sequence + '.svg'
        tDs = RNA.svg_rna_plot(sequence, structure, output_file)

        return tDs
    
    def stick_and_dots(self,sequence,PAM_sequence, gRNA_tail_sequence): #создание модели шпилек gRNA
        sequence = sequence + PAM_sequence + gRNA_tail_sequence
        fc  = RNA.fold_compound(sequence)
        ss = fc.mfe()

        return ss

    def image_placer(self,path_to_file, file_to_place,path_to_pngs, column_to_place): #добавление png картинок с вторичной структурой gRNA в excel
        workbook = openpyxl.load_workbook(path_to_file + file_to_place + '.xlsx')
        ws = workbook.active
        for row in range(2,ws.max_row + 1):
            ws.row_dimensions[row].height = 100
            index_value = ws[f'A{row}'].value
            image_filename = f'{path_to_pngs}{index_value}.png'
            img = Image(image_filename)
            img.width = 100 #размеры картинок
            img.height = 100
            cell_to_add = f'{column_to_place}{row}'
            ws.add_image(img,cell_to_add)
            workbook.save(path_to_file + 'with_images_'+ file_to_place + '.xlsx')

    def csv_to_dict(self,path_to_csv):
        dict_res = {}
        with open(path_to_csv, 'r',encoding='utf-8') as file:
            comma = csv.reader(file)
            for i in comma:
                key = i[0]
                value = float(i[1])
                dict_res[key] = value
        return dict_res


    def start_code_by_button(self):
        self.main_part()

    if not os.path.exists('download/'):
        os.makedirs('download/',exist_ok=True)   

    def main_part(self):     
        self.directory_to_input = 'download/'
        self.path_to_input_file = self.to_input_file_frame.get()

        
        self.file_name = os.path.basename(self.path_to_input_file)
        self.file_name = os.path.splitext(self.file_name)[0]
        print(f'self.file_name: {self.file_name}')
        self.main_directory_to_export = self.file_name + '/'
        print(f'self.main_directory_to_export: {self.main_directory_to_export}')

        self.default_configs = {
            'min_gc': 40,
            'max_gc': 70,
            'oligoT': 666,
            'forv_strain': 0,
            'rev_strain': 1,
            'GC_freq': 0,
            'GC_count_10to20': 0,
            'last_four_pur_1': 0.25,
            'last_four_pur_2': 0.5,
            'last_four_pur_3': 0.75,
            'last_four_pur_4': 1,
            'access_to_20': 2,
            'access_to_19': 1,
            'access_to_18': 1,
            'seven_links': 666,
            'twelve_links': 666,
            'C_not_at_3': 0.5,
            'G_not_at_16': 0.5,
            'C_at_16': 0.5,
            'G_or_A_at_20': 0.5,
            'C_at_18': 0.5,
            'G_not_at_14': 0.5,
            'GCC_not_at_16to20': 666,
            'PAMs_A': 0,
            'PAMs_T': -1,
            'PAMs_G': 1,
            'PAMs_C': 1
            }

        if not os.path.exists('configs.csv'):
            with open('configs.csv', 'w', newline='', encoding='utf-8') as file:
                cfg = csv.writer(file)
                for key, value in self.default_configs.items():
                    cfg.writerow([key, value])       

        if self.configs_path.get() == '':
            self.configs = self.default_configs
        else:
            self.configs = self.csv_to_dict(self.configs_path.get())

        if not os.path.exists(self.main_directory_to_export):
            os.makedirs(self.main_directory_to_export,exist_ok=True)
        if not os.path.exists(self.main_directory_to_export + '/png'):
            os.makedirs(self.main_directory_to_export + '/png',exist_ok=True)

        if  self.path_to_input_file == '': #для разработки. если ничего не задавать в поле названия файла, то откроется 'файл по умолчанию'
            self.path = self.directory_to_input + 'guides_2xLoxPafterCre_pz9Athaliana-unknownLoc.xls'
        else:
            self.path =  self.path_to_input_file

        if self.input_RNA_tail_seq_frame.get() == '':
            self.gRNA_tail = 'GTTTCAGAGCTATGCTGGAAACAGCATAGCAAGTTGAAATAAGGCTAGTCCGTTATCAACTTGAAAAAGTGGCACCGAGTCGGTGCTTTT'
        else:
            self.gRNA_tail = self.input_RNA_tail_seq_frame.get()

        print(f'Result saved at: {self.path}')
        print(f'Chosen gRNA tail: {self.gRNA_tail}')
        self.raw_seq_DF = pd.read_excel(self.path, skiprows=8)
        pd.set_option('display.max_rows', None)
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        
        self.sequences_list = list()
        for i, row in self.raw_seq_DF.iterrows():
            self.current_gRNA_w_PAM = row['targetSeq']
            self.current_gRNA_wo_PAM = row['targetSeq'][0:20]
            self.current_PAM = self.current_gRNA_w_PAM[20:24]

            self.current_gRNA_seq = gRNA_sequence(sequence=self.current_gRNA_wo_PAM,
                                                  PAM= self.current_PAM,
                                                  GC_count_10to20= self.GC_structure(self.current_gRNA_wo_PAM)[0],
                                                  GC_freq=self.GC_structure(self.current_gRNA_wo_PAM)[1],
                                                  last_four_pur=self.GC_structure(self.current_gRNA_wo_PAM)[2])
            
            if str(row['#guideId']).endswith('forw'):
                self.current_gRNA_seq.strain = self.configs['forv_strain']
                self.current_gRNA_seq.total_score += self.configs['forv_strain']
            elif str(row['#guideId']).endswith('rev'):
                self.current_gRNA_seq.strain = self.configs['rev_strain']
                self.current_gRNA_seq.total_score += self.configs['rev_strain']                

            if 'TTTT' in  self.current_gRNA_wo_PAM:
                self.current_gRNA_seq.oligoT = 666

            if self.current_gRNA_seq.last_four_pur == 1:
                self.current_gRNA_seq.total_score += self.configs['last_four_pur_1']
            elif self.current_gRNA_seq.last_four_pur == 2:
                self.current_gRNA_seq.total_score += self.configs['last_four_pur_2']
            elif self.current_gRNA_seq.last_four_pur == 3:
                self.current_gRNA_seq.total_score += self.configs['last_four_pur_3']
            elif self.current_gRNA_seq.last_four_pur == 4:
                self.current_gRNA_seq.total_score += self.configs['last_four_pur_4']

            if self.current_gRNA_wo_PAM[2:3] != 'C': #цитозин не желателен в 3 положении
                self.current_gRNA_seq.C_not_at_3 = self.configs['C_not_at_3']   
                self.current_gRNA_seq.total_score += self.configs['C_not_at_3'] 

            if self.current_gRNA_wo_PAM[15:16] != 'G': #гуанин нежелателен в 16 положении
                self.current_gRNA_seq.G_not_at_16 = self.configs['G_not_at_16'] 
                self.current_gRNA_seq.total_score += self.configs['G_not_at_16']  

            if self.current_gRNA_wo_PAM[15:16] == 'C': #в 16 положении желателен цитозин
                self.current_gRNA_seq.C_at_16 = self.configs['C_at_16'] 
                self.current_gRNA_seq.total_score += self.configs['C_at_16']  

            if self.current_gRNA_wo_PAM[19:20] == 'G' or self.current_gRNA_wo_PAM[19:20] == 'A': #аденин или гуанин желателен в 20 положении
                self.current_gRNA_seq.G_or_A_at_20 = self.configs['G_or_A_at_20']
                self.current_gRNA_seq.total_score += self.configs['G_or_A_at_20'] 

            if self.current_gRNA_wo_PAM[17:18] == 'C': #цитозин желателен в 18 положении
                self.current_gRNA_seq.C_at_18 = self.configs['C_at_18'] 
                self.current_gRNA_seq.total_score += self.configs['C_at_18']  

            if self.current_gRNA_wo_PAM[13:14] != 'G': #гуанин не желателен в 14 положении
                self.current_gRNA_seq.G_not_at_14 = self.configs['G_not_at_14']
                self.current_gRNA_seq.total_score += self.configs['G_not_at_14'] 

            if 'GCC' in self.current_gRNA_wo_PAM[15:20] : #GCC не должно быть на участке 16-20
                self.current_gRNA_seq.GCC_not_at_16to20 = self.configs['GCC_not_at_16to20']

            if self.current_PAM[0:1] == 'G': #в PAM на месте N желателен гуанин или цитозин и нежелателен тимин
                self.current_gRNA_seq.PAMs_N = self.configs['PAMs_G']
                self.current_gRNA_seq.total_score += self.configs['PAMs_G']
            elif self.current_PAM[0:1] == 'C': 
                self.current_gRNA_seq.PAMs_N = self.configs['PAMs_C']
                self.current_gRNA_seq.total_score += self.configs['PAMs_C']    
            elif self.current_PAM[0:1] == 'T': 
                self.current_gRNA_seq.PAMs_N = self.configs['PAMs_T']
                self.current_gRNA_seq.total_score += self.configs['PAMs_T']   
            elif self.current_PAM[0:1] == 'A': 
                self.current_gRNA_seq.PAMs_N = self.configs['PAMs_A']
                self.current_gRNA_seq.total_score += self.configs['PAMs_A']            

            self.sticks_structure = self.stick_and_dots(self.current_gRNA_wo_PAM, self.current_PAM, self.gRNA_tail)#создание модели шпильки

            if self.sticks_structure[0][19] == '(' or self.sticks_structure[0][19] == ')': #20 нуклеотид очень желательно не должен быть связан
                self.current_gRNA_seq.access_18_to_20 += 0
            else:
                self.current_gRNA_seq.access_18_to_20 += self.configs['access_to_20'] 
                self.current_gRNA_seq.total_score += self.configs['access_to_20']             

            if self.sticks_structure[0][18] == '(' or self.sticks_structure[0][18] == ')':#19 нуклеотид желательно не должен быть связан
                self.current_gRNA_seq.access_18_to_20 += 0
            else:
                self.current_gRNA_seq.access_18_to_20 += self.configs['access_to_19']
                self.current_gRNA_seq.total_score += self.configs['access_to_19'] 
            if self.sticks_structure[0][17] == '(' or self.sticks_structure[0][17] == ')':#18 нуклеотид желательно не должен быть связан
                self.current_gRNA_seq.access_18_to_20 += 0
            else:
                self.current_gRNA_seq.access_18_to_20 += self.configs['access_to_18']
                self.current_gRNA_seq.total_score += self.configs['access_to_18'] 

            print(f'{i}\n{self.sticks_structure[0][0:20]}\n{self.current_gRNA_seq.sequence}')
            if '(((((((' in self.sticks_structure[0][0:20]: #проверка нет ли 7 связанных подряд
                self.current_gRNA_seq.seven_links += self.configs['seven_links']

            if self.sticks_structure[0][0:20].count('(') + self.sticks_structure[0][0:20].count(')') >=12: #проверка нет ли 12 связанных вообще
                self.current_gRNA_seq.twelve_links += self.configs['twelve_links']  

            '''
            следующая игра с удалением толькочто созданных svg файлов - костыль
            '''
            self.RNA_2D_structure(self.current_gRNA_wo_PAM, self.current_PAM, self.gRNA_tail)
            cairosvg.svg2png(url= self.current_gRNA_wo_PAM + self.current_PAM + self.gRNA_tail + '.svg',write_to= self.main_directory_to_export + '/png/' + str(i)+ '.png')
            time.sleep(0.1)#сон 0.1 сек             
            os.remove(self.current_gRNA_wo_PAM + self.current_PAM + self.gRNA_tail + '.svg') #удаление тысяч svg файлов


            self.sequences_list.append(self.current_gRNA_seq) #расширение списка с классами последовательностей gRNA





        self.output_excel = pd.DataFrame([i for i in self.sequences_list])#создание из списка pdDataFrame
        self.dirty_excel = self.output_excel
        self.output_excel.to_excel(self.main_directory_to_export + 'not_clear_'+  self.file_name + '.xlsx')#создаём xlsx файл без выкидываний gRNA

        self.output_excel = self.output_excel[self.output_excel['GC_freq'] >= self.default_configs['min_gc']] #выкидываем с неподходящим GC составом
        self.output_excel = self.output_excel[self.output_excel['GC_freq'] <= self.default_configs['max_gc']]
        self.output_excel = self.output_excel[self.output_excel['GCC_not_at_16to20'].isin([0])] #выброс последовательностей с GCC на последних 5 нуклеотидах
        self.output_excel = self.output_excel[self.output_excel['oligoT'].isin([0])] #выброс последовательностей с олиго T (>4) в составе
        self.output_excel = self.output_excel[self.output_excel['seven_links'].isin([0])] #выброс последовательностей с более чем 7 связанынми подряд
        self.output_excel = self.output_excel[self.output_excel['twelve_links'].isin([0])] #выброс последовательностей с 12 и более связанными вообще

        self.output_excel = self.output_excel.drop('GCC_not_at_16to20',axis=1)
        self.output_excel = self.output_excel.drop('oligoT',axis=1)
        self.output_excel = self.output_excel.drop('seven_links',axis=1)
        self.output_excel = self.output_excel.drop('twelve_links',axis=1)

        '''
        следующая игра с созданием-удалением файлов - костыль
        '''
        self.output_excel.to_excel(self.main_directory_to_export + 'clear_'+  self.file_name + '.xlsx')
        self.image_placer(self.main_directory_to_export, 'clear_'+ self.file_name,self.main_directory_to_export + '/png/','R')
        self.image_placer(self.main_directory_to_export, 'not_clear_'+ self.file_name,self.main_directory_to_export + '/png/','V') #в грязный файл добавляем картинки
        time.sleep(1) #спим 1 секунду для удаления файлов без картинок
        os.remove(self.main_directory_to_export + 'clear_'+  self.file_name + '.xlsx')
        os.remove(self.main_directory_to_export + 'not_clear_'+  self.file_name + '.xlsx')
        self.programm_is_done(self.main_directory_to_export)

        
if __name__ == '__main__':
    app = App()

    app.mainloop()

